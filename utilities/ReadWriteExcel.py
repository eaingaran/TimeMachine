import logging
import os

from openpyxl import load_workbook, Workbook

from utilities import Path


def read_cell(book_name, sheet_name, cell):
    work_book = None
    try:
        work_book = load_workbook(Path.get_base_path() + book_name)
        sheet = work_book[sheet_name]
        value = sheet[cell].value
        logging.debug("Values of {} in sheet {} of book {} is {}".format(cell, sheet_name, book_name, value))
        work_book.close()
        return value
    except Exception as e:
        logging.error("Couldn't get the value fo the cell {} in sheet {} of book {}" \
                      .format(cell, sheet_name, book_name), e)
        if work_book is not None:
            work_book.close()
        return None


def write_cell(book_name, sheet_name, cell, value):
    work_book = None
    try:
        work_book = load_workbook(Path.get_base_path() + book_name, read_only=False, keep_vba=True)
        sheet = work_book[sheet_name]
        sheet[cell] = value
        logging.debug("Values of {} in sheet {} of book {} is set to {}".format(cell, sheet_name, book_name, value))
        work_book.save(Path.get_base_path() + book_name)
        work_book.close()
    except FileNotFoundError as e:
        logging.error("File {} is not found.".format(book_name), e)
        create_excel_workbook(book_name, sheet_name)
        try:
            write_cell(book_name, sheet_name, cell, value)
        except Exception as e:
            logging.error("Tried creating a new sheet and calling the method again. Still failed.", e)
    except Exception as e:
        logging.error("Error while trying to write.", e)
        if work_book is not None:
            work_book.close()


def create_excel_workbook(book_name, sheet_name):
    work_book = None
    try:
        os.makedirs(os.path.dirname(Path.get_base_path() + book_name), exist_ok=True)
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = sheet_name
        work_book.save(filename=Path.get_base_path() + book_name)
        work_book.close()
        logging.debug("Work book with name {} with a sheet named {} created.".format(book_name, sheet_name))
    except Exception as e:
        logging.error("Couldn't create a workbook.", e)
        if work_book is not None:
            work_book.close()
